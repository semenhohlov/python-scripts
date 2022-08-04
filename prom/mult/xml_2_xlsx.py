from xml.dom.minidom import parse
import os
import re
import openpyxl
import datetime
import json
import mysql.connector

def load_config():
    try:
        f = open('config.json', 'r')
        config = json.load(f)
        f.close()
        print('Загрузка файла конфигурации "config.json"')
        return config
    except:
        print('Ошибка загрузки файла конфигурации "config.json"')
        return False

def rename_old_price():
    if os.path.exists('export_mult.xlsx'):
        os.rename('export_mult.xlsx', 'prev.xlsx')
        print('Переименовую предыдущий export_mult.xlsx в prev.xlsx.')

def get_dom(name):
    dom = False
    try:
        dom = parse(name)
    except:
        print('Ошибка чтения файла: ', name)
    return dom

def get_categories(dom):
    cats = {}
    categories = dom.getElementsByTagName('category')
    for cat in categories:
        try:
            id = int(cat.getAttribute('id'))
            value = str(cat.childNodes[0].nodeValue)
            cats[id] = value
        except:
            pass
    return cats

def get_supplier(dom):
    name = dom.getElementsByTagName('name')[0]
    return str(name.childNodes[0].nodeValue)

def load_accordance(supplier=''):
    if not supplier:
        print('Не указан поставщик для выбора групп.')
        return
    mydb = mysql.connector.connect(
        host='localhost',
        user='admin',
        password='123',
        database='prom'
    )
    cursor = mydb.cursor()
    cursor.execute(
        '''
        select accordance_categories.group_number,
        accordance_categories.group_id,
        accordance_categories.group_name,
        accordance_categories.group_prom_name,
        categories.prom_group_id,
        categories.prom_cat_adress
        from accordance_categories 
        left join (categories)
        on (accordance_categories.group_number = categories.group_number)
        where supplier = "{}"
        '''.format(supplier))
    tmp = cursor.fetchall()
    result = {}
    for t in tmp:
        result[t[1]] = t
    cursor.close()
    mydb.close()
    return result

def meas(str):
    tmp = str.split(' х ')
    if len(tmp) == 1:
        tmp.append('')
        tmp.append('')
    if len(tmp) == 2:
        tmp.append('')
    return tmp

def get_char_column(work_sheet, char_number=0):
    a1 = 'Название_Характеристики'
    b1 = 'Измерение_Характеристики'
    c1 = 'Значение_Характеристики'
    last_column = 49
    pos = char_number * 3 + last_column + 1
    if not work_sheet.cell(1, pos).value:
        work_sheet.cell(1, pos).value = a1
        work_sheet.cell(1, pos + 1).value = b1
        work_sheet.cell(1, pos + 2).value = c1
    return pos

def insert_row(item, work_sheet, row, config, cats):
    row_str = str(row)
    desc_begin = '<p>Магазин Никусик порадует вас огромным ассортиментом по доступным ценам! </p><p>'
    desc_end = '</p><p> В нашем магазине обновление товаров - несколько раз в неделю, и вы обязательно подберёте для себя что-нибудь подходящее! А если не знаете, что купить на подарок - звоните или пишите, мы обязательно вам поможем сделать удачный выбор!</p>'

    # kod_tovara
    try:
        vendorCode =  item.getElementsByTagName('vendorCode')[0]
        work_sheet["A" + row_str] = str(vendorCode.childNodes[0].nodeValue)
    except:
        pass
    
    # pos_name
    nameItem = item.getElementsByTagName('name')[0]
    name = str(nameItem.childNodes[0].nodeValue).replace('"', '')
    name = name.replace('р/у', 'на радиоуправлении')
    name = name.replace('р.у.', 'на радиоуправлении')
    name = name.replace('&qout;', '')
    name = name.replace('&amp;qout;', '')
    name = name.replace('БТ Холодильник', 'Игрушка Холодильник')
    name = re.sub("\(\d*шт\)", "", name)
    work_sheet["B" + row_str] = name
    
    # description
    desc =  item.getElementsByTagName('description')[0]
    value = ''
    try:
        value = str(desc.childNodes[0].nodeValue)
    except:
        pass
    work_sheet["F" + row_str] = desc_begin + value + desc_end
    
    # price
    price =  item.getElementsByTagName('price')[0]
    value = float(price.childNodes[0].nodeValue)
    work_sheet["I" + row_str] = round(value * config['over_price'])
    
    # UAH
    work_sheet["J" + row_str] = "UAH"
    
    # шт.
    work_sheet["K" + row_str] = "шт."
    
    # image_url
    try:
        img =  item.getElementsByTagName('picture')[0]
        work_sheet["O" + row_str] = str(img.childNodes[0].nodeValue)
    except:
        pass
    
    # present
    work_sheet["P" + row_str] = '+'
    # work_sheet["P" + row_str] = source_ws["S" + row_str].value
    
    # kol-vo na ostatke
    work_sheet["Q" + row_str] = 1
    
    # group_id
    group_id =  item.getElementsByTagName('categoryId')[0]
    gr_id = int(group_id.childNodes[0].nodeValue)
    try:
        acc = cats[gr_id]
        work_sheet["R" + row_str] = acc[0]
        # group_name
        work_sheet["S" + row_str] = acc[3]
    except:
        print('Группы прома не совпадают с группами загружаемого прайса.')
        print('Запустите сравнение групп.')
        exit()
    try:
        # prom_cat_adress
        work_sheet["T" + row_str] = acc[5]
        
        # prom_cat_id
        work_sheet["Z" + row_str] = acc[4]
    except:
        print('Категории прома не совпадают с группами.')
        print('Запустите сравнение категорий.')
        exit()
    
    # id
    work_sheet["Y" + row_str] = config['prefix'] + str(item.getAttribute('id'))
    
    # brend
    # vendor =  item.getElementsByTagName('vendor')[0]
    # work_sheet["AB" + row_str] = str(vendor.childNodes[0].nodeValue)
    
    # discount
    work_sheet["AD" + row_str] = config['discount_value']
    
    # discount_begin
    work_sheet["AH" + row_str] = config['discount_begin']
    
    # discount_end
    work_sheet["AI" + row_str] = config['discount_end']
    
    # personal marks
    work_sheet["AF" + row_str] = "Multitoys"
    # yarlik
    work_sheet["AK" + row_str] = "Склад 4"
    
    # weight
    # work_sheet["AR" + row_str] = '0'
    
    # sostoyanie
    # work_sheet["AX" + row_str] = "Состояние"
    
    # novoe
    # work_sheet["AZ" + row_str] = "Новое"

    params = item.getElementsByTagName('param')
    char_number = 0
    for param in params:
        if param.getAttribute('name') == 'Страна':
            work_sheet['AC' + row_str].value = param.childNodes[0].nodeValue
        if param.getAttribute('name') == 'Бренд':
            work_sheet['AB' + row_str].value = param.childNodes[0].nodeValue
        if param.getAttribute('name') == 'Упаковка':
            work_sheet['W' + row_str].value = param.childNodes[0].nodeValue
        if param.getAttribute('name') == 'Размер упаковки':
            gabs = meas(str(param.childNodes[0].nodeValue))
            # length
            work_sheet["AU" + row_str] = gabs[0]
            # width
            work_sheet["AS" + row_str] = gabs[1]
            # height
            work_sheet["AT" + row_str] = gabs[2]
        pos = get_char_column(work_sheet, char_number)
        work_sheet.cell(row, pos).value = param.getAttribute('name')
        work_sheet.cell(row, pos + 2).value = param.childNodes[0].nodeValue
        char_number += 1
            

def make_new_price(work_sheet, xml, config):
    row = 2
    cats = get_categories(xml)
    print('Всего категорий {}.'.format(len(cats)))
    acc = load_accordance(get_supplier(xml))
    print('загружено категорий {}.'.format(len(acc)))
    itemList = xml.getElementsByTagName('offer')
    total = len(itemList)
    percent = int(total / 100)
    current = 1
    tmp = 0
    print('Всего позиций {}.'.format(total))
    for item in itemList:
        insert_row(item, work_sheet, row, config, acc)
        row += 1
        tmp += 1
        if tmp > percent:
            tmp = 0
            current += 1
            print('\r{}%'.format(current), end='')
    print()

def append_row(work_sheet, row):
    last_row = work_sheet.max_row + 1
    current_cell = 1
    for cell in row:
        if current_cell == 16:
            work_sheet.cell(last_row, current_cell).value = '-'
        else:
            work_sheet.cell(last_row, current_cell).value = cell.value
        current_cell += 1

def add_no_presents_items(work_sheet):
    print('Добавление позиций которых нет в наличии.')
    if not os.path.exists('prev.xlsx'):
        print('Файл prev.xlsx отсутствует.')
        return
    prev = openpyxl.load_workbook("prev.xlsx")
    prev_ws = prev.active
    new_items = []
    for row in work_sheet.rows:
        new_items.append(row[0].value)
    diff = 0
    total = prev_ws.max_row
    percent = int(total / 100)
    current = 1
    tmp = 0
    for row in prev_ws.rows:
        try:
            new_items.index(row[0].value)
        except:
            if row[15].value == '+':
                append_row(work_sheet, row)
                diff += 1
        tmp += 1
        if tmp > percent:
            tmp = 0
            current += 1
            print('\r{}%'.format(current), end='')
    print()
    print('Нет в наличии {} позиций'.format(diff))

    prev.close()

def main():
    dt = datetime.datetime.now()
    print('Время начала: {}:{}.'.format(dt.hour, dt.minute))
    config = load_config()
    temp = openpyxl.load_workbook("template.xlsx")
    work_sheet = temp.active
    # print('Colunms', work_sheet.max_column)
    # print('Rows', work_sheet.max_row)
    # print('1:49', work_sheet.cell(1, 49).value)
    # return
    xml_file_name = 'mult.xml'
    print('Загрузки xml файла "{}"'.format(xml_file_name))
    xml = get_dom(xml_file_name)
    if not xml:
        print('Ошибка загрузки {}'.format(xml_file_name))
        return
    print('xml файл "{}" загружен.'.format(xml_file_name))
    rename_old_price()
    make_new_price(work_sheet, xml, config)
    # add_no_presents_items(work_sheet)
    temp.save('export_mult.xlsx')
    temp.close()
    dt = datetime.datetime.now()
    print('xml файл обработан: {}:{}.'.format(dt.hour, dt.minute))

if __name__ == '__main__':
    main()